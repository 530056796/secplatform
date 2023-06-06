# -*- coding:UTF-8 -*-

import os
import sys
import datetime
import requests
import urllib3
import time
import json
import re
import random
import string
import shutil
import openpyxl
import traceback
from SenInfo import models

class scan_test3:

    def __init__(self,**kwargs):
        self.start_time = kwargs['starttime']
        self.end_time = kwargs['endtime']
        self.cftk = kwargs['cftk']
        self.Auth_Token = kwargs['Auth_Token']
        self.regex_str = kwargs['regex_str']
        self.url = kwargs['Murl']
        self.taskid = kwargs['taskid']
        self.finishtime = kwargs['finishtime']
        self.appname = kwargs['appname']
        self.appid = kwargs['appid']

    def t_split(self):
        # str转时间格式
        start_time_switch = datetime.datetime.strptime(self.start_time, "%Y-%m-%d %H:%M:%S")
        end_time_switch = datetime.datetime.strptime(self.end_time, "%Y-%m-%d %H:%M:%S")
        end_time_z = (end_time_switch + datetime.timedelta(hours=-8)).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3]+"Z"
        print(end_time_z)

        # 生成间隔一分钟的时间列表
        t0 = start_time_switch
        if not self.appname and not self.appid or "bos" in self.appname and not self.appid or "hrmp" in self.appname and not self.appid:
            interval = 1
        else:
            interval = 10

        timelist = []
        while t0 < end_time_switch:
            t0_z = (t0 + datetime.timedelta(hours=-8)).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3]+"Z"
            #print(t0_z)
            timelist.append(t0_z)
            #print(timelist)
            #print(len(timelist))
            t0 = t0 + datetime.timedelta(seconds=interval)
        timelist.append(end_time_z)
        #print(timelist)
        print(len(timelist))
        return timelist


    def report(self):

        # 创建扫描报告:以当前时间作为名称复制模板，生成报告
        report_str ="report_" + time.strftime("%Y%m%d%H%M%S",time.localtime()) + ''.join(random.sample(string.ascii_letters + string.digits, 8)) + ".xlsx"
        reportfile_path = os.getcwd().replace('\\', '/') + "/static/report/" + report_str
        shutil.copy("report_template.xlsx", reportfile_path)
        return reportfile_path

    def mscan(self):
        try:
            reportfile_path = scan_test3.report(self)

            # 加载打开报告中的report表页签
            workbook = openpyxl.load_workbook(filename=reportfile_path)
            sheet_report = workbook['report']

            #时间拆分
            timelist = scan_test3.t_split(self)
            print(timelist)

            #隔10/1S，循环
            x = 0
            t = len(timelist) - 2

            while x <= t:
                timelist_min = timelist[x:x+2]

            #请求获取日志信息
                requests.packages.urllib3.disable_warnings()

                #url = 'https://feature.kingdee.com:1026/feature_sit_hr/monitor/eye/logQuery'

                payload = {"appName":self.appname,"appId":self.appid,"instanceId":"","traceId":"","className":"","methodName":"","opKey":"","opMethod":"",\
                           "formId":"","formName":"","userID":"","userName":"","tenantId":"","level":"",\
                           "time":timelist_min,"keyword":"","slowQuery":"false","size":10000,"ip":""}


                headers = {'Host':'devhr-test-386367.kcssz.cloud.kingdee.com',\
                           'monitor-csrf-token':self.cftk,\
                           'Content-Type':'application/json;charset=UTF-8',\
                           'Cookie':'Auth_Token='+self.Auth_Token}

                res = requests.post(self.url,verify=False,json=payload,headers=headers)
                #print(res.status)
                #print(res.text)

                #json字符串转为字典格式
                body = json.loads(res.text)
                print ("轮次:"+str(x)+",日志条数:"+str(len(body['data'])),timelist_min)

                #regex_str = self.regex_str
                pattern = re.compile(self.regex_str,re.I)

                m = 0
                n = sheet_report.max_row
                print("当前行数："+str(n))

                while m < len(body['data']):
                    # print (body['data'][m]['appName'],body['data'][m]['className'],body['data'][m]['methodName'],body['data'][m]['line'],body['data'][m]['message'],body['data'][m]['instanceId'],body['data'][m]['time'])

                    # print("===================================================================")
                    # print(m)
                    # print(body['data'][m]['message'])
                    # print(body['data'][m])
                    if 'message' in body['data'][m]:
                        match_key = pattern.findall(re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', r'', body['data'][m]['message']))
                        match_key = list(set(match_key))
                        #print(match_key)

                        if not match_key:
                            pass
                        else:
                            sheet_report.cell(row=n + 1, column=1).value = n
                            sheet_report.cell(row=n + 1, column=2).value = body['data'][m]['appName']
                            if 'className' in body['data'][m]:
                                sheet_report.cell(row=n + 1, column=3).value = body['data'][m]['className']
                            else:
                                pass
                            sheet_report.cell(row=n + 1, column=4).value = body['data'][m]['methodName']

                            if 'line' in body['data'][m]:
                                sheet_report.cell(row=n + 1, column=5).value = body['data'][m]['line']
                            else:
                                pass

                            sheet_report.cell(row=n + 1, column=6).value = ",".join(match_key)

                            sheet_report.cell(row=n + 1, column=7).value = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', r'', body['data'][m]['message'])

                            sheet_report.cell(row=n + 1, column=7).data_type = 's'

                            if 'line' in body['data'][m]:
                                sheet_report.cell(row=n + 1, column=8).value = body['data'][m]['instanceId']
                            else:
                                pass

                            sheet_report.cell(row=n + 1, column=10).value = body['data'][m]['time']

                            if 'appId' not in body['data'][m]['logtags']:
                                pass
                            else:
                                #print(body['data'][m]['logtags'])
                                sheet_report.cell(row=n + 1, column=9).value = body['data'][m]['logtags']['appId']
                            n = n + 1
                    else:
                        pass

                    m = m + 1

                #统计扫描进度
                if (x == t):
                    pro_percent = '100%'
                elif (t >= 2000) and (x%20 == 0):
                    p1 = 100 * x//t
                    pro_percent = f'{p1}%'
                elif (800< t < 2000) and (x % 15 == 0):
                    p1 = 100 * x // t
                    pro_percent = f'{p1}%'
                elif (100< t <= 800) and (x % 8 == 0):
                    p1 = 100 * x // t
                    pro_percent = f'{p1}%'
                elif (t <= 100) :
                    p1 = 100 * x // t
                    pro_percent = f'{p1}%'

                print(pro_percent)

                models.Scaninfo.objects.filter(taskid=self.taskid).update(progress=pro_percent)

                x += 1

            workbook.save(filename=reportfile_path)

            models.Scaninfo.objects.filter(taskid=self.taskid).update(finishtime=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
            models.Scaninfo.objects.filter(taskid=self.taskid).update(report=reportfile_path.split("/")[-1])

        except Exception as e:


            loglist = []
            loglist.append("【"+time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())+"】")
            loglist.append(" 【任务id】："+self.taskid+";")
            logfile_path = os.getcwd().replace('\\', '/') + "/error_log/" + time.strftime("%Y%m%d",time.localtime()) + ".log"

            with open(logfile_path, "a+") as f:
                loglist.append(" 【ERROR】："+traceback.format_exc())
                f.writelines(loglist)
                print(traceback.format_exc())
            #错误日志补充

            models.Scaninfo.objects.filter(taskid=self.taskid).update(progress='Scan Error!')

            pass



if __name__ == "__main__":
    ticks1 = time.time()
    param = {"starttime": "2022-10-21 14:00:00", "endtime": "2022-10-21 14:00:04", "Murl": "https://feature.kingdee.com:2024/baseline_a/monitor/eye/logQuery", "Auth_Token": "Du5cjrBQNoCp4624BrD9PWWIOVq7sj51", "cftk": "FLYQkKNfBWhZiNmjHQf43mXDEtDSOzVE", "regex_str": ".", "taskname": "test215", "taskbegintime": "2022-10-21 14:39:09", "finishtime": "", "progress": "", "report": "", "id": "9XxWwdCMv5sRV8Au"}
    scan1 = scan_test3(**param)
    scan1.mscan()
    ticks2 = time.time()
    print(ticks2-ticks1)
